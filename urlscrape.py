#!/usr/bin/python3
import requests
from requests.adapters import HTTPAdapter
from requests.exceptions import ConnectionError, Timeout
import openpyxl
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from getpass import getpass
from time import sleep
import sys
import datetime
from datetime import date
import urllib
from urllib.parse import urlparse
import logging


import http.client as http_client
#http_client.HTTPConnection.debuglevel = 1

def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '|'):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print('\r%s |%s| %s%% %s' % (prefix, bar, percent, suffix), end = '\r')
    # Print New Line on Complete
    if iteration == total: 
        print()


def main():
	#logging.basicConfig()
	#logging.getLogger().setLevel(logging.DEBUG)
	#requests_log = logging.getLogger("requests.packages.urllib3")
	#requests_log.setLevel(logging.DEBUG)
	#requests_log.propagate = True
	todaystring = date.today().strftime("%Y-%m-%d")
	now = datetime.datetime.now()
	midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
	seconds = (now - midnight).seconds

	wb = load_workbook("daytest.xlsx")
	ws = wb.active
	username = str(input("username: "))
	password = getpass()
	agencyURL = str(input("Agency subdomain: "))
	#batch_number = str(input("Batch #: "))
	
	with requests.Session() as sess:
		awards_adapter = HTTPAdapter(max_retries=3)
		sess.mount("https://" + agencyURL +".footholdtechnology.com", awards_adapter)
		login_payload = {'handle': username , 'password': password} 
		login = sess.post("https://" + agencyURL + ".footholdtechnology.com/zf2/", data=login_payload)

		if "loggedInUser" in login.text:
			print("Logged in!")
		else:
			print("Hmm, I couldnt log you in")
			sys.exit()
		sleep(2)
		printProgressBar(0, ws.max_row-1, prefix = 'Progress:', suffix = 'Complete', length = 50)
		for row in ws.iter_rows(min_row=2,max_col=16):
			if row[13].value == None:
				invoice_number_payload = {
					'ser_no': str(row[2].value), # Must create the batch FIRST
					'pagemode': 'Continue Add',
					'clientid_INT': str(row[4].value),
					#'progname_char': '+'.join(row[2].value.split(" ")), # How to handle other characters? urllib.parse.quote(row[2].value) To-do: %20 or + for spaces?
					#'progname_char': urllib.parse.quote(row[2].value), # How to handle other characters? urllib.parse.quote(row[2].value) To-do: %20 or + for spaces?
					'progname_char': row[3].value, # Content-Type: application/x-www-form-urlencoded encodes everything on it's own.  
					'continue': 'CONTINUE',
					'loginname_char': username,
					'codefile_CHAR': 'bientry.php',
					'pagedate': todaystring,
					'pagetime': str(seconds),
				}
				# Get the invoice number and NPI
				try:
					r_invoice_number = sess.post("https://" + agencyURL + ".footholdtechnology.com/bientry.php?urltime=" + str(seconds), data=invoice_number_payload,timeout=3)
				except ConnectionError as ce:
					row[13].value = str(ce)
					wb.save(filename="daytest.xlsx")
					continue

				except Timeout:
					row[13].value = str("Timed out")
					wb.save(filename="daytest.xlsx")
					continue
				except Exception as err:
					row[13].value = str(f'Another error occurred {err}')
					wb.save(filename="daytest.xlsx")
					continue
				try:
					soup = BeautifulSoup(r_invoice_number.content,'html.parser')
					newInvoiceNumber = soup.find("input", {"name":"ch-inv_no[]"})['value']
					npi = soup.find("input", {"name":"ch-npi[]"})["value"]
				except:
					print("UH-OH! Something weird happened.")
					row[13].value = "Error at line 81.  Moving on to the next row of data."
					continue
				sleep(2)
				addInvoicePayload = {
					'inv_dt': row[6].value.strftime('%m/%d/%Y'),
					'ch-inv_no[]': newInvoiceNumber,
					'ser_no': str(row[2].value),
					'clientid_int': str(row[4].value),
					'ch-inv_line[]': '1',
					'ch-diagnosis[]': str(row[5].value),
					'ch-svc_dt[]': row[6].value.strftime('%m/%d/%Y'),
					'ch-proccode[]': str(row[7].value),
					'ch-npi[]': npi,
					'ch-times[]': str(row[9].value),
					'ch-amount[]': str(row[10].value),
					'mcadjust[]': 'A',
					'origcrn[]': str(row[11].value),
					'authorization_ids[]': str(row[12].value),
					'pagemode': 'ADD THIS INVOICE',
					#'progname_char': '+'.join(row[2].value.split(" ")),
					#'progname_char': urllib.parse.quote(row[2].value),
					'progname_char': row[3].value, # Content-Type: application/x-www-form-urlencoded encodes everything on it's own.  
					'loginname_char': username,
					'codefile_CHAR': 'bientry.php',
					'pagedate': todaystring,
					'pagetime': str(seconds),
				}

				try:
					r_add_invoice = sess.post("https://" + agencyURL + ".footholdtechnology.com/bientry.php?urltime=" + str(seconds), data=addInvoicePayload)
				except ConnectionError as ce:
					row[13].value = str(ce)
					wb.save(filename="daytest.xlsx")
					continue
				except Timeout:
					row[13].value = str("Timed out")
					wb.save(filename="daytest.xlsx")
					continue
				except Exception as err:
					row[13].value = str(f'Another error occurred: {err}')
					wb.save(filename="daytest.xlsx")
					continue
				if "Add Invoice" in r_add_invoice.text:
					now = datetime.datetime.now()
					nowTs = now.strftime(";%m/%d/%Y, %H:%M:%S")
					row[13].value = str(newInvoiceNumber+nowTs)
					wb.save(filename="daytest.xlsx") # Save the file after each row? i/o intense?
				else:
					row[13].value = "I did not find add invoice! :( Will move on to the next row..."
				seconds+=1 # Increment seconds so we don't timeout from server
			printProgressBar(row[0].row-1, ws.max_row-1, prefix = 'Progress:', suffix = 'Complete', length = 50)
	wb.save(filename="daytest.xlsx")		
if __name__ == "__main__":
    main()
